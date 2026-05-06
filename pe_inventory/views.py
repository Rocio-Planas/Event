import json
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views import View
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError

from .models import Item, StandAssignment
from in_person_events.models import Event


@method_decorator(login_required, name='dispatch')
class InventoryDashboardView(TemplateView):
    """
    Vista principal del dashboard de inventario.
    RF-06: Control de Recursos Básicos.
    
    Inyecta todos los datos necesarios para el frontend.
    """
    template_name = 'pe_inventory/inventory.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        event_id = self.kwargs.get('event_id')
        
        # Obtener el evento
        event = get_object_or_404(Event, id=event_id)
        
        # Obtener todos los items con propiedades calculadas
        items = Item.objects.all()
        
        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'name': item.name,
                'category': item.category,
                'total_stock': item.total_stock,
                'used_stock': item.used_stock,
                'available_stock': item.available_stock,
                'status': item.status,
                'image': item.image.url if item.image else None,
                'notes': item.notes,
            })
        
        # Calcular estadísticas globales usando los valores de status
        total_in_stock = sum(1 for i in items_data if i['status'] == 'En Stock')
        total_low_stock = sum(1 for i in items_data if i['status'] == 'Stock Bajo')
        total_no_stock = sum(1 for i in items_data if i['status'] == 'Sin Stock')
        
        context['items_json'] = json.dumps(items_data)
        context['items'] = items
        context['total_items'] = len(items_data)
        context['total_in_stock'] = total_in_stock
        context['total_low_stock'] = total_low_stock
        context['total_no_stock'] = total_no_stock
        context['event'] = event
        context['event_id'] = event_id
        context['active_page'] = 'inventario'
        
        return context


@login_required
def create_item(request, event_id):
    """
    API para crear un nuevo recurso.
    POST /inventario/<event_id>/create/
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Validar campos requeridos
        name = request.POST.get('name', '').strip()
        category = request.POST.get('category', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        if not name:
            return JsonResponse({'error': 'El nombre del artículo es requerido'}, status=400)
        if not category:
            return JsonResponse({'error': 'La categoría es requerida'}, status=400)
        # Validar stock total
        try:
            total_stock = int(request.POST.get('total_stock', 0))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'El stock total debe ser un número válido'}, status=400)
        
        if total_stock < 1:
            return JsonResponse({'error': 'El stock total debe ser mayor a 0'}, status=400)
        
        item = Item(
            name=name,
            category=category,
            total_stock=total_stock,
            image=request.FILES.get('image') if request.FILES else None,
            notes=notes,
        )
        item.full_clean()
        item.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Recurso creado correctamente',
            'item': {
                'id': item.id,
                'name': item.name,
                'category': item.category,
                'total_stock': item.total_stock,
                'used_stock': item.used_stock,
                'available_stock': item.available_stock,
                'status': item.status,
                'image': item.image.url if item.image else None,
                'notes': item.notes,
            }
        })
    
    except ValidationError as e:
        return JsonResponse({'error': 'Error de validación: ' + str(e.message if hasattr(e, 'message') else e)}, status=400)
    except IntegrityError as e:
        return JsonResponse({'error': 'Error de integridad de datos: ' + str(e)}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al crear el recurso: {str(e)}'}, status=500)


@login_required
def update_item(request, event_id, item_id):
    """
    API para actualizar un recurso existente.
    POST /inventario/<event_id>/update/<item_id>/
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    item = get_object_or_404(Item, id=item_id)
    
    try:
        if 'name' in request.POST:
            name = request.POST['name'].strip()
            if not name:
                return JsonResponse({'error': 'El nombre no puede estar vacío'}, status=400)
            item.name = name
        
        if 'category' in request.POST:
            category = request.POST['category'].strip()
            if not category:
                return JsonResponse({'error': 'La categoría no puede estar vacía'}, status=400)
            item.category = category
        
        if 'total_stock' in request.POST:
            try:
                total_stock = int(request.POST['total_stock'])
            except (ValueError, TypeError):
                return JsonResponse({'error': 'El stock total debe ser un número válido'}, status=400)
            if total_stock < 1:
                return JsonResponse({'error': 'El stock total debe ser mayor a 0'}, status=400)
            item.total_stock = total_stock
        
        if 'notes' in request.POST:
            item.notes = request.POST['notes'].strip()
        
        if 'image' in request.FILES:
            item.image = request.FILES['image']
        
        item.full_clean()
        item.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Recurso actualizado correctamente',
            'item': {
                'id': item.id,
                'name': item.name,
                'category': item.category,
                'total_stock': item.total_stock,
                'used_stock': item.used_stock,
                'available_stock': item.available_stock,
                'status': item.status,
                'image': item.image.url if item.image else None,
                'notes': item.notes,
            }
        })
    
    except ValidationError as e:
        return JsonResponse({'error': 'Error de validación: ' + str(e.message if hasattr(e, 'message') else e)}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al actualizar el recurso: {str(e)}'}, status=500)


@login_required
def delete_item(request, event_id, item_id):
    """
    API para eliminar un recurso.
    POST /inventario/<event_id>/delete/<item_id>/
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    item = get_object_or_404(Item, id=item_id)
    
    try:
        item_name = item.name
        item.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Recurso "{item_name}" eliminado correctamente'
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al eliminar el recurso: {str(e)}'}, status=500)


@login_required
def get_items(request, event_id):
    """
    API para obtener todos los recursos.
    GET /inventario/<event_id>/items/
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        items = Item.objects.all().order_by('name')
        items_data = []
        
        for item in items:
            items_data.append({
                'id': item.id,
                'name': item.name,
                'category': item.category,
                'total_stock': item.total_stock,
                'used_stock': item.used_stock,
                'available_stock': item.available_stock,
                'status': item.status,
                'image': item.image.url if item.image else None,
                'notes': item.notes,
            })
        
        return JsonResponse({
            'success': True,
            'items': items_data
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al obtener los recursos: {str(e)}'}, status=500)


@login_required
def export_excel(request, event_id):
    """
    Exporta todos los items del inventario a un archivo Excel.
    GET /inventario/<event_id>/export-excel/
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Definir estilos
        header_fill = PatternFill(start_color="0058BE", end_color="0058BE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Categoría', 'Stock Total', 'Stock Usado', 'Stock Disponible', 'Estado', 'Notas']
        ws.append(headers)
        
        # Aplicar estilos al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Obtener items
        items = Item.objects.all().order_by('name')
        
        # Agregar datos
        for item in items:
            ws.append([
                item.id,
                item.name,
                item.category,
                item.total_stock,
                item.used_stock,
                item.available_stock,
                item.status,
                item.notes,
            ])
        
        # Aplicar estilos a las celdas de datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al exportar: {str(e)}'}, status=500)


@login_required
def import_excel(request, event_id):
    """
    Importa items del archivo Excel al inventario.
    Soporta crear nuevos items o actualizar existentes.
    POST /inventario/<event_id>/import-excel/
    """
    from openpyxl import load_workbook
    from io import BytesIO
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No se proporcionó archivo'}, status=400)
    
    try:
        file = request.FILES['file']
        
        # Validar extensión
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}, status=400)
        
        # Cargar workbook
        file_content = BytesIO(file.read())
        wb = load_workbook(file_content)
        ws = wb.active
        
        # Validar encabezados
        expected_headers = ['ID', 'Nombre', 'Categoría', 'Stock Total', 'Stock Usado', 'Stock Disponible', 'Estado', 'Notas']
        actual_headers = [cell.value for cell in ws[1]]
        
        if actual_headers[:len(expected_headers)] != expected_headers:
            return JsonResponse({
                'error': 'El formato del archivo no es válido. Los encabezados deben ser: ' + ', '.join(expected_headers)
            }, status=400)
        
        imported_count = 0
        updated_count = 0
        errors = []
        warnings = []
        
        # Procesar filas
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                item_id = row[0].value
                name = row[1].value
                category = row[2].value
                total_stock = row[3].value
                notes = row[7].value if row[7] else ''
                
                # Validaciones
                if not name:
                    errors.append(f'Fila {row_idx}: El nombre es requerido')
                    continue
                
                if not category:
                    errors.append(f'Fila {row_idx}: La categoría es requerida')
                    continue
                
                try:
                    total_stock = int(total_stock) if total_stock else 1
                except (ValueError, TypeError):
                    errors.append(f'Fila {row_idx}: El stock total debe ser un número')
                    continue
                
                if total_stock < 1:
                    errors.append(f'Fila {row_idx}: El stock total debe ser mayor a 0')
                    continue
                
                # Validar categoría
                valid_categories = [choice[0] for choice in Item.Category.choices]
                if category not in valid_categories:
                    errors.append(f'Fila {row_idx}: Categoría "{category}" no válida')
                    continue
                
                # Crear o actualizar
                if item_id:
                    try:
                        item_id = int(item_id)
                    except (ValueError, TypeError):
                        item_id = None
                
                if item_id:
                    try:
                        item = Item.objects.get(id=item_id)
                        item.name = name
                        item.category = category
                        item.total_stock = total_stock
                        item.notes = notes
                        item.save()
                        updated_count += 1
                    except Item.DoesNotExist:
                        item = Item(
                            name=name,
                            category=category,
                            total_stock=total_stock,
                            notes=notes,
                        )
                        item.full_clean()
                        item.save()
                        imported_count += 1
                        warnings.append(f'Fila {row_idx}: El item con ID {row[0].value} no existe, se creó un nuevo registro')
                else:
                    item = Item(
                        name=name,
                        category=category,
                        total_stock=total_stock,
                        notes=notes,
                    )
                    item.full_clean()
                    item.save()
                    imported_count += 1
            
            except Exception as e:
                errors.append(f'Fila {row_idx}: Error - {str(e)}')
        
        response_data = {
            'success': True,
            'message': f'Importación completada: {imported_count} nuevos items, {updated_count} items actualizados',
            'imported': imported_count,
            'updated': updated_count,
        }
        if errors:
            response_data['errors'] = errors
        if warnings:
            response_data['warnings'] = warnings

        return JsonResponse(response_data)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al importar: {str(e)}'}, status=500)